from flask import Flask, request, render_template, redirect, url_for, send_file
import pandas as pd
import os
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
from reportlab.lib import colors
from reportlab.pdfgen import canvas
import io
import webbrowser
import time
from threading import Thread, Lock
import signal
import glob
import atexit
import socket
import urllib.request
import subprocess
import sys

# Caminhos
base_path = os.path.dirname(os.path.abspath(__file__))
logo_path = os.path.join(base_path, 'static', 'images', 'logo.png')

# Sinalizar ambiente de nuvem (Azure App Service) via variáveis conhecidas
IS_CLOUD = bool(
    os.getenv("WEBSITE_SITE_NAME")
    or os.getenv("WEBSITE_INSTANCE_ID")
    or os.getenv("PORT")
)
PORT = int(os.environ.get("PORT", 5000))

# Variáveis globais para controle de timeout
last_activity_time = time.time()
activity_lock = Lock()
timeout_minutes = 3
shutdown_initiated = False


# Função para atualizar atividade (deve estar antes do middleware)
def update_activity():
    """Atualiza o timestamp da última atividade"""
    global last_activity_time
    with activity_lock:
        last_activity_time = time.time()

# Configurações do Flask
app = Flask(__name__)
app.config['UPLOAD_FOLDER'] = 'uploads'

# Criar pasta para uploads, se não existir
if not os.path.exists(app.config['UPLOAD_FOLDER']):
    os.makedirs(app.config['UPLOAD_FOLDER'])

# Middleware para atualizar atividade em todas as requisições
@app.before_request
def before_request():
    """Atualiza timestamp de atividade antes de cada requisição"""
    update_activity()

@app.route('/')
def index():
    """Página inicial que exibe o formulário de upload"""
    return render_template('index.html')

# Rota de saúde para monitoramento pelo Azure/App Service
@app.route('/healthz')
def healthz():
    return {"status": "ok"}, 200

@app.route('/upload', methods=['POST'])
def upload_file():
    """Recebe o arquivo enviado pelo usuário e processa"""
    if 'file' not in request.files:
        return "Erro: Nenhum arquivo enviado", 400

    file = request.files['file']

    if file.filename == '':
        return "Erro: Nenhum arquivo selecionado", 400

    if file:
        filename = file.filename
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        file.save(filepath)
        
        # Verificar qual modo de leitura usar
        reading_mode = detect_reading_mode(filepath)
        return redirect(url_for('resultado', filename=filename, mode=reading_mode))

def detect_reading_mode(filepath):
    """Detecta automaticamente qual modo de leitura usar baseado nas colunas do arquivo"""
    try:
        print(f"🔍 Analisando arquivo: {filepath}")
        
        import openpyxl
        wb = openpyxl.load_workbook(filepath)
        ws = wb.active
        
        print(f"📊 Arquivo tem {ws.max_row} linhas")
        
        # Procurar pelo formato específico da planilha do datalogger (id, Data/Hora, Temperatura, Umidade)
        for row_idx in range(1, min(50, ws.max_row + 1)):
            row_values = [str(cell.value).strip().lower() if cell.value else '' for cell in ws[row_idx]]
            
            # Debug: mostrar algumas linhas
            if row_idx <= 15:
                print(f"Linha {row_idx}: {row_values[:5]}...")  # Primeiras 5 colunas
            
            # Verificar se é o formato específico do datalogger (id, Data/Hora, Temperatura[°C], Umidade[%Hr])
            if ('id' in row_values and 
                any('data' in val and 'hora' in val for val in row_values) and
                any('temperatura' in val for val in row_values) and
                any('umidade' in val for val in row_values)):
                wb.close()
                print(f"✅ Modo datalogger detectado na linha {row_idx}")
                return 'new_mode'
        
        # Procurar pelo formato de relatório nas primeiras 50 linhas
        for row_idx in range(1, min(50, ws.max_row + 1)):
            row_values = [str(cell.value).strip().lower() if cell.value else '' for cell in ws[row_idx]]
            
            # Verificar se é o formato de relatório (N°., Temp, UR, Tempo)
            if ('n°.' in row_values or 'nº.' in row_values or 'n°' in row_values) and \
               ('temp' in row_values) and \
               ('ur' in row_values) and \
               ('tempo' in row_values):
                wb.close()
                print(f"✅ Modo relatório detectado na linha {row_idx}")
                return 'report_mode'
        
        # Verificar se há indicadores de relatório no início do arquivo
        print("🔍 Procurando indicadores de relatório...")
        for row_idx in range(1, min(20, ws.max_row + 1)):
            for cell in ws[row_idx]:
                if cell.value and isinstance(cell.value, str):
                    cell_text = str(cell.value).strip().lower()
                    if 'relatório' in cell_text:
                        print(f"📋 Encontrado 'relatório' na linha {row_idx}")
                    if 'início:' in cell_text:
                        print(f"⏰ Encontrado 'início:' na linha {row_idx}")
                    if 'tax amostr.:' in cell_text:
                        print(f"📊 Encontrado 'tax amostr.:' na linha {row_idx}")
                    
                    if 'relatório' in cell_text or \
                       ('início:' in cell_text and 'fim:' in cell_text) or \
                       'tax amostr.:' in cell_text or \
                       'dado n°.:' in cell_text:
                        wb.close()
                        print(f"✅ Indicadores de relatório detectados na linha {row_idx}")
                        return 'report_mode'
        
        wb.close()
        print("❌ Formato de relatório não detectado, testando outros modos...")
        
        # Tentar diferentes valores de skiprows para o formato original
        for skip_rows in [3, 2, 1, 0, 4, 5]:
            try:
                df = pd.read_excel(filepath, skiprows=skip_rows)
                if df.empty or len(df.columns) < 3:
                    continue
                    
                # Normalizar as colunas para minúsculas e remover espaços
                df.columns = df.columns.astype(str).str.strip().str.lower()
                
                print(f"Testando modo original com skiprows {skip_rows}: {list(df.columns)}")
                
                # Verificar padrões de colunas para o modo atual (original)
                current_mode_patterns = [
                    ['sn', 'date', 'time', 'oc', '%rh'],  # Formato completo
                    ['date', 'oc', '%rh'],  # Formato básico
                    ['date', 'time', 'oc', '%rh'],  # Com time separado
                    ['sn', 'date', 'oc', '%rh']  # Com SN mas sem time
                ]
                
                for pattern in current_mode_patterns:
                    if all(col in df.columns for col in pattern):
                        print(f"Modo original detectado com padrão: {pattern}")
                        return 'current_mode'
                
                # Verificar por palavras-chave do formato original
                col_str = ' '.join(df.columns).lower()
                if (any(keyword in col_str for keyword in ['date', 'data']) and
                    any(keyword in col_str for keyword in ['oc', 'temp']) and
                    any(keyword in col_str for keyword in ['%rh', 'rh', 'humid'])):
                    print("Modo original detectado por palavras-chave")
                    return 'current_mode'
                    
            except Exception as e:
                print(f"Erro ao testar skiprows={skip_rows}: {e}")
                continue
        
        # Se chegou até aqui, usar modo atual como padrão
        print("Usando modo original como padrão")
        return 'current_mode'
        
    except Exception as e:
        print(f"Erro na detecção: {e}")
        return 'current_mode'

def process_report_mode(filepath, filename):
    """Processa o arquivo usando o modo de relatório"""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(filepath)
        ws = wb.active
        
        # Encontrar a linha dos cabeçalhos da tabela de dados
        header_row = None
        for row_idx in range(1, min(50, ws.max_row + 1)):
            row_values = [str(cell.value).strip().lower() if cell.value else '' for cell in ws[row_idx]]
            
            # Procurar pela linha com N°., Temp, UR, Tempo
            if ('n°.' in row_values or 'nº.' in row_values or 'n°' in row_values) and \
               ('temp' in row_values) and \
               ('ur' in row_values) and \
               ('tempo' in row_values):
                header_row = row_idx
                break
        
        if header_row is None:
            raise Exception("Não foi possível encontrar os cabeçalhos da tabela de dados")
        
        wb.close()
        
        # Ler o arquivo Excel a partir da linha dos cabeçalhos
        skip_rows = header_row - 1
        df = pd.read_excel(filepath, skiprows=skip_rows)
        
        if df.empty:
            raise Exception("Arquivo vazio após aplicar skiprows")
        
        # Normalizar as colunas para minúsculas
        df.columns = df.columns.astype(str).str.strip().str.lower()
        
        print(f"Processando modo relatório com skiprows={skip_rows}, colunas: {list(df.columns)}")
        
        # Mapear as colunas (ignorando N°.)
        temp_col = None
        ur_col = None
        tempo_col = None
        
        for col in df.columns:
            if 'temp' in col and not 'tempo' in col:
                temp_col = col
            elif 'ur' in col:
                ur_col = col
            elif 'tempo' in col:
                tempo_col = col
        
        if not all([temp_col, ur_col, tempo_col]):
            raise Exception(f"Colunas necessárias não encontradas. Encontradas: {list(df.columns)}")
        
        # Filtrar apenas as colunas necessárias (ignorando N°.)
        df_filtered = df[[tempo_col, temp_col, ur_col]].copy()
        
        # Renomear para padrão
        df_filtered.columns = ['tempo', 'temperatura', 'umidade']
        
        # Converter a coluna 'tempo' para datetime
        df_filtered['tempo'] = pd.to_datetime(df_filtered['tempo'], errors='coerce')
        
        # CORREÇÃO: Converter colunas de temperatura e umidade para numérico
        # Remover caracteres não numéricos e converter para float
        df_filtered['temperatura'] = pd.to_numeric(
            df_filtered['temperatura'].astype(str).str.replace(r'[^0-9.,\-]', '', regex=True).str.replace(',', '.'),
            errors='coerce'
        )
        
        df_filtered['umidade'] = pd.to_numeric(
            df_filtered['umidade'].astype(str).str.replace(r'[^0-9.,\-]', '', regex=True).str.replace(',', '.'),
            errors='coerce'
        )
        
        # Remover linhas com dados inválidos
        df_filtered = df_filtered.dropna(subset=['tempo', 'temperatura', 'umidade'])
        
        if df_filtered.empty:
            raise Exception("Nenhum dado válido encontrado após conversão")
        
        # Extrair apenas a data (ignorando a hora)
        df_filtered['data'] = df_filtered['tempo'].dt.date
        
        # Calcular os valores máximos e mínimos por dia
        result = df_filtered.groupby('data').agg(
            Temp_Max=('temperatura', 'max'),
            Temp_Min=('temperatura', 'min'),
            Umid_Max=('umidade', 'max'),
            Umid_Min=('umidade', 'min')
        ).reset_index()
        
        # Formatar a coluna 'data' no formato de data brasileiro (DD/MM/YYYY)
        result['data'] = result['data'].apply(lambda x: x.strftime('%d/%m/%Y'))
        
        # Ajustar a formatação dos números
        result['Temp_Max'] = result['Temp_Max'].round(2)
        result['Temp_Min'] = result['Temp_Min'].round(2)
        result['Umid_Max'] = result['Umid_Max'].round(2)
        result['Umid_Min'] = result['Umid_Min'].round(2)
        
        # Renomear as colunas para exibir os nomes personalizados
        result = result.rename(columns={
            'data': 'Data',
            'Temp_Max': 'Temperatura Máxima (°C)',
            'Temp_Min': 'Temperatura Mínima (°C)',
            'Umid_Max': 'Umidade Máxima (%)',
            'Umid_Min': 'Umidade Mínima (%)'
        })
        
        # Salvar o dataframe como variável global para uso na geração do PDF
        global latest_result
        latest_result = result
        
        # Converter a tabela para HTML e remover quebras de linha extras
        table_html = result.to_html(classes='table table-striped table-bordered', index=False)
        table_html = table_html.replace("\n", "")
        
        # Exibir resultados
        return render_template(
            'resultado.html',
            table=table_html,
            filename=filename
        )
        
    except Exception as e:
        print(f"Erro no processamento do modo relatório: {e}")
        return (
            f"Erro: Não foi possível processar o arquivo. {str(e)}",
            400,
        )

@app.route('/resultado/<filename>', methods=['GET', 'POST'])
def resultado(filename):
    """Processa o arquivo Excel e exibe os resultados filtrados"""
    filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
    reading_mode = request.args.get('mode', 'current_mode')
    
    try:
        if reading_mode == 'report_mode':
            return process_report_mode(filepath, filename)
        elif reading_mode == 'new_mode':
            return process_new_mode(filepath, filename)
        else:
            return process_current_mode(filepath, filename)
    except Exception as e:
        return f"Erro ao processar o arquivo: {e}", 500

def process_current_mode(filepath, filename):
    """Processa o arquivo usando o modo de leitura atual (original)"""
    # Tentar diferentes configurações de skiprows
    for skip_rows in [3, 2, 1, 0, 4, 5]:
        try:
            df = pd.read_excel(filepath, skiprows=skip_rows)
            if df.empty:
                continue
                
            # Normalizar as colunas para minúsculas
            df.columns = df.columns.astype(str).str.strip().str.lower()
            
            print(f"Tentando modo original com skiprows={skip_rows}, colunas: {list(df.columns)}")
            
            # Tentar diferentes mapeamentos de colunas
            column_mappings = [
                # Formato completo com SN, DATE, TIME, oC, %RH
                {'date_col': 'date', 'temp_col': 'oc', 'humid_col': '%rh', 'time_col': 'time'},
                # Formato básico DATE, oC, %RH
                {'date_col': 'date', 'temp_col': 'oc', 'humid_col': '%rh'},
                # Variações de nomes
                {'date_col': 'date', 'temp_col': 'temp', 'humid_col': 'rh'},
                {'date_col': 'data', 'temp_col': 'temperatura', 'humid_col': 'umidade'}
            ]
            
            for mapping in column_mappings:
                # Verificar se as colunas essenciais existem
                required_cols = [mapping['date_col'], mapping['temp_col'], mapping['humid_col']]
                if all(col in df.columns for col in required_cols):
                    
                    # Se há coluna TIME separada, combinar DATE e TIME
                    if 'time_col' in mapping and mapping['time_col'] in df.columns:
                        # Combinar data e hora
                        df['datetime'] = pd.to_datetime(
                            df[mapping['date_col']].astype(str) + ' ' + df[mapping['time_col']].astype(str),
                            errors='coerce'
                        )
                    else:
                        # Usar apenas a coluna de data
                        df['datetime'] = pd.to_datetime(df[mapping['date_col']], errors='coerce')
                    
                    # Remover linhas com datas inválidas
                    df = df.dropna(subset=['datetime'])
                    
                    if df.empty:
                        continue
                    
                    # Extrair apenas a data (ignorando a hora)
                    df['date'] = df['datetime'].dt.date
                    
                    # Renomear colunas para facilitar o processamento
                    df = df.rename(columns={
                        mapping['temp_col']: 'temperatura',
                        mapping['humid_col']: 'umidade'
                    })
                    
                    # Calcular os valores máximos e mínimos por dia
                    result = df.groupby('date').agg(
                        Temp_Max=('temperatura', 'max'),
                        Temp_Min=('temperatura', 'min'),
                        Umid_Max=('umidade', 'max'),
                        Umid_Min=('umidade', 'min')
                    ).reset_index()
                    
                    # Formatar a coluna 'date' no formato de data brasileiro (DD/MM/YYYY)
                    result['date'] = result['date'].apply(lambda x: x.strftime('%d/%m/%Y'))
                    
                    # Ajustar a formatação dos números
                    result['Temp_Max'] = result['Temp_Max'].round(2)
                    result['Temp_Min'] = result['Temp_Min'].round(2)
                    result['Umid_Max'] = result['Umid_Max'].round(2)
                    result['Umid_Min'] = result['Umid_Min'].round(2)
                    
                    # Renomear as colunas para exibir os nomes personalizados
                    result = result.rename(columns={
                        'date': 'Data',
                        'Temp_Max': 'Temperatura Máxima (°C)',
                        'Temp_Min': 'Temperatura Mínima (°C)',
                        'Umid_Max': 'Umidade Máxima (%)',
                        'Umid_Min': 'Umidade Mínima (%)'
                    })
                    
                    # Salvar o dataframe como variável global para uso na geração do PDF
                    global latest_result
                    latest_result = result
                    
                    # Converter a tabela para HTML e remover quebras de linha extras
                    table_html = result.to_html(classes='table table-striped table-bordered', index=False)
                    table_html = table_html.replace("\n", "")
                    
                    # Exibir resultados
                    return render_template(
                        'resultado.html',
                        table=table_html,
                        filename=filename
                    )
                    
        except Exception as e:
            print(f"Erro com skiprows={skip_rows}: {e}")
            continue
    
    # Se chegou até aqui, não conseguiu processar
    return (
        f"Erro: Não foi possível identificar as colunas necessárias no arquivo. "
        f"Verifique se o arquivo contém colunas de data, temperatura e umidade.",
        400,
    )

def process_new_mode(filepath, filename):
    """Processa o arquivo usando o novo modo de leitura (formato datalogger)"""
    try:
        # Encontrar a linha correta dos cabeçalhos
        import openpyxl
        wb = openpyxl.load_workbook(filepath)
        ws = wb.active
        
        header_row = None
        for row_idx in range(1, min(50, ws.max_row + 1)):
            row_values = [str(cell.value).strip().lower() if cell.value else '' for cell in ws[row_idx]]
            
            # Procurar pela linha com id, Data/Hora, Temperatura, Umidade
            if ('id' in row_values and 
                any('data' in val and 'hora' in val for val in row_values) and
                any('temperatura' in val for val in row_values) and
                any('umidade' in val for val in row_values)):
                header_row = row_idx
                print(f"📍 Cabeçalhos encontrados na linha {row_idx}")
                break
        
        wb.close()
        
        if header_row is None:
            raise Exception("Não foi possível encontrar os cabeçalhos corretos (id, Data/Hora, Temperatura, Umidade)")
        
        # Ler o arquivo com o skiprows correto
        skip_rows = header_row - 1
        df = pd.read_excel(filepath, skiprows=skip_rows)
        
        if df.empty:
            raise Exception("Arquivo vazio após aplicar skiprows")
            
        # Normalizar as colunas para minúsculas e remover espaços extras
        df.columns = df.columns.astype(str).str.strip().str.lower()
        
        print(f"🔄 Processando modo datalogger com skiprows={skip_rows}")
        print(f"📋 Colunas encontradas: {list(df.columns)}")
        
        # Mapear as colunas encontradas de forma mais flexível
        id_col = None
        datetime_col = None
        temp_col = None
        humid_col = None
        
        for col in df.columns:
            col_clean = col.strip().lower()
            if col_clean == 'id':
                id_col = col
            elif 'data' in col_clean and 'hora' in col_clean:
                datetime_col = col
            elif 'temperatura' in col_clean:
                temp_col = col
            elif 'umidade' in col_clean:
                humid_col = col
        
        print(f"🗂️ Mapeamento de colunas:")
        print(f"   ID: {id_col}")
        print(f"   Data/Hora: {datetime_col}")
        print(f"   Temperatura: {temp_col}")
        print(f"   Umidade: {humid_col}")
        
        if not all([datetime_col, temp_col, humid_col]):
            raise Exception(f"Colunas necessárias não encontradas. Encontradas: {list(df.columns)}")
        
        # Filtrar as colunas necessárias (ignorando o campo 'id')
        df_filtered = df[[datetime_col, temp_col, humid_col]].copy()
        
        # Renomear para padrão
        df_filtered.columns = ['data_hora', 'temperatura', 'umidade']
        
        print(f"📊 Dados antes da conversão: {len(df_filtered)} linhas")
        
        # Converter a coluna 'data_hora' para datetime
        df_filtered['data_hora'] = pd.to_datetime(df_filtered['data_hora'], errors='coerce')
        
        # Converter colunas de temperatura e umidade para numérico
        # Remover caracteres não numéricos e converter para float
        df_filtered['temperatura'] = pd.to_numeric(
            df_filtered['temperatura'].astype(str).str.replace(r'[^0-9.,\-]', '', regex=True).str.replace(',', '.'),
            errors='coerce'
        )
        
        df_filtered['umidade'] = pd.to_numeric(
            df_filtered['umidade'].astype(str).str.replace(r'[^0-9.,\-]', '', regex=True).str.replace(',', '.'),
            errors='coerce'
        )
        
        # Remover linhas com dados inválidos
        df_filtered = df_filtered.dropna(subset=['data_hora', 'temperatura', 'umidade'])
        
        if df_filtered.empty:
            raise Exception("Nenhum dado válido encontrado após conversão")
        
        print(f"✅ Dados válidos após conversão: {len(df_filtered)} linhas")
        
        # Extrair apenas a data (ignorando a hora)
        df_filtered['data'] = df_filtered['data_hora'].dt.date
        
        # Calcular os valores máximos e mínimos por dia
        result = df_filtered.groupby('data').agg(
            Temp_Max=('temperatura', 'max'),
            Temp_Min=('temperatura', 'min'),
            Umid_Max=('umidade', 'max'),
            Umid_Min=('umidade', 'min')
        ).reset_index()
        
        # Formatar a coluna 'data' no formato de data brasileiro (DD/MM/YYYY)
        result['data'] = result['data'].apply(lambda x: x.strftime('%d/%m/%Y'))
        
        # Ajustar a formatação dos números
        result['Temp_Max'] = result['Temp_Max'].round(2)
        result['Temp_Min'] = result['Temp_Min'].round(2)
        result['Umid_Max'] = result['Umid_Max'].round(2)
        result['Umid_Min'] = result['Umid_Min'].round(2)
        
        # Renomear as colunas para exibir os nomes personalizados
        result = result.rename(columns={
            'data': 'Data',
            'Temp_Max': 'Temperatura Máxima (°C)',
            'Temp_Min': 'Temperatura Mínima (°C)',
            'Umid_Max': 'Umidade Máxima (%)',
            'Umid_Min': 'Umidade Mínima (%)'
        })
        
        # Salvar o dataframe como variável global para uso na geração do PDF
        global latest_result
        latest_result = result
        
        # Converter a tabela para HTML e remover quebras de linha extras
        table_html = result.to_html(classes='table table-striped table-bordered', index=False)
        table_html = table_html.replace("\n", "")
        
        # Exibir resultados
        return render_template(
            'resultado.html',
            table=table_html,
            filename=filename
        )
        
    except Exception as e:
        print(f"Erro no processamento do novo modo: {e}")
        return (
            f"Erro: Não foi possível processar o arquivo. {str(e)}",
            400,
        )

@app.route('/gerar_pdf', methods=['GET', 'POST'])
def gerar_pdf():
    """Gera um PDF com os resultados filtrados"""
    try:
        # Receber os parâmetros do formulário
        param1 = request.form.get('param1')  # Formulação
        param2 = request.form.get('param2')  # Revisão
        param3 = request.form.get('param3')  # 'aprovado' ou 'reprovado'
        param4 = request.form.get('param4')  # Data fornecida no formulário
        param5 = request.form.get('param5')  # Número do estudo
        param6 = request.form.get('param6')  # Código do equipamento
        param7 = request.form.get('param7')  # Número do ensaio
        param8 = request.form.get('param8')  # Local de leitura do equipamento

        # Para voltar a ser dinamico remova os 'disable' dos html e apague da linha 129 a 132
        param1 = "FOR.2.031"  # Formulação
        param2 = "Rev. 00"  # Revisão
        param3 = "Aprovado"  # 'aprovado' ou 'reprovado'

        # Garantir que os valores são strings
        param1 = str(param1) if param1 else "oi1"
        param2 = str(param2) if param2 else "oi2"
        param3 = str(param3) if param3 else "oi3"
        param4 = str(param4) if param4 else "24/03/2025"
        param5 = str(param5) if param5 else "oi5"
        param6 = str(param6) if param6 else "oi6"
        param7 = str(param7) if param7 else "oi7"
        param8 = str(param8) if param8 else "oi8"

        # Verifica se os resultados estão disponíveis
        if 'latest_result' not in globals() or latest_result.empty:
            return "Erro: Nenhum dado disponível para gerar o PDF.", 400

        # Configura o PDF
        buffer = io.BytesIO()

        # Calcular o número total de páginas
        items_per_page = 30  # Defina o número de itens por página
        total_items = len(latest_result)  # Número total de linhas (excluindo cabeçalhos)
        total_pages = (total_items // items_per_page) + (1 if total_items % items_per_page > 0 else 0)


        def add_header(canvas, doc, is_first_page, total_pages, param1, param2, param3, param4, param5, param6, param7, param8):
            """Função para adicionar o cabeçalho"""
            canvas.saveState()

            # Cabeçalho principal
            canvas.rect(9, 720, 592, 60)
            import sys
            import os

            def get_resource_path(relative_path):
                """Retorna o caminho correto do recurso, seja no ambiente normal ou no .exe."""
                if getattr(sys, 'frozen', False):  # Se estiver rodando no PyInstaller (.exe)
                    base_path = sys._MEIPASS
                else:
                   if getattr(sys, 'frozen', False):  # Se estiver rodando num executável (.exe)
                      base_path = sys._MEIPASS
                   else:
                    base_path = os.path.dirname(os.path.abspath(sys.argv[0]))  # Usa o primeiro argumento ao invés de __file__
                
                return os.path.join(base_path, relative_path)

            logo_path = get_resource_path("static/images/logo.png")


            try:
                canvas.drawImage(logo_path, 35, 730, width=100, height=40, mask='auto')
            except Exception as e:
                raise Exception(f"Erro ao carregar a logo: {e}")
            canvas.line(150, 780, 150, 720)
            canvas.line(450, 780, 450, 720)
            canvas.line(450, 750, 601, 750) #linha horizontal do meio canto direito
            canvas.line(523, 780, 523, 720) #linha vertical do meio canto direito

            # Título e informações
            canvas.setFont("Helvetica-Bold", 13)
            canvas.drawString(160, 745, "DADOS DE TEMPERATURA E/OU UMIDADE")
            canvas.setFont("Helvetica", 10)
            canvas.drawString(455, 760, f"{param1}")  # Formulação
            canvas.setFont("Helvetica", 10)
            canvas.drawString(550, 760, f"{doc.page} / {total_pages}")  # Paginas
            canvas.drawString(470, 730, f"{param2}")  # Revisao
            canvas.drawString(540, 737, f"{param3}")  # 'aprovado' ou 'reprovado'
            canvas.drawString(535, 725, f"{param4}")  # Data

            # Tabela adicional (agora em todas as páginas)
            data = [["Número do estudo:", str(param5), "Código do equipamento:", str(param6)],
                    ["Número do ensaio:", str(param7), "Local de leitura do equipamento:", str(param8)]]
            table = Table(data, colWidths=[100,169,155,169], rowHeights=20)
            table.setStyle(TableStyle([
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ]))

            table.wrapOn(canvas, 9, 670)
            table.drawOn(canvas, 9, 670)

            canvas.restoreState()

        def add_footer(canvas, doc):
            """Função para adicionar o rodapé com o campo de assinatura"""
            canvas.saveState()
            canvas.setFont("Helvetica", 10)
            canvas.drawString(40, 50, "Rubrica: ________________________________________________")
            canvas.drawString(40, 30, "Data: ___________________________________________________")
            canvas.restoreState()

        elements = []

        # Dados para a tabela principal
        data = [latest_result.columns.tolist()]
        data += latest_result.values.tolist()

        # Configuração da tabela
        table = Table(data, colWidths=[102, 123, 123, 123, 123], repeatRows=1)
        style = TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.white),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ])
        table.setStyle(style)
        elements.append(table)

        doc = SimpleDocTemplate(buffer, pagesize=letter, topMargin=126)
        doc.build(
            elements,
            onFirstPage=lambda c, d: (add_header(c, d, True, total_pages, param1, param2, param3, param4, param5, param6, param7, param8), add_footer(c, d)),
            onLaterPages=lambda c, d: (add_header(c, d, False, total_pages, param1, param2, param3, param4, param5, param6, param7, param8), add_footer(c, d))
        )

        buffer.seek(0)
        return send_file(
            buffer,
            mimetype='application/pdf',
            as_attachment=False,
            download_name='Resultado Final.pdf'
        )

    except Exception as e:
        return f"Erro ao gerar o PDF: {e}", 500

@app.route('/cleanup-uploads', methods=['POST'])
def cleanup_uploads_route():
    """Rota para limpar uploads quando solicitado pelo frontend"""
    try:
        cleanup_uploads()
        return '', 204  # No Content
    except Exception as e:
        print(f"Erro na limpeza via rota: {e}")
        return '', 500

def close_browser():
    """Fecha todas as abas do navegador relacionadas à aplicação"""
    try:
        if sys.platform == "win32":
            # Windows - fecha janelas do Chrome/Edge
            subprocess.run([
                'powershell', '-Command',
                'Get-Process | Where-Object {$_.ProcessName -eq "chrome" -or $_.ProcessName -eq "msedge"} | ForEach-Object { $_.CloseMainWindow() }'
            ], capture_output=True, timeout=5)
        elif sys.platform == "darwin":
            # macOS - fecha abas do Chrome com localhost (aproximação)
            subprocess.run([
                'osascript', '-e',
                'tell application "Google Chrome" to close (tabs of windows whose URL contains "localhost:5000")'
            ], capture_output=True, timeout=5)
        elif sys.platform.startswith("linux"):
            # Linux - encerra processos relacionados
            subprocess.run(['pkill', '-f', 'localhost:5000'], capture_output=True, timeout=5)
    except Exception as e:
        print(f"Aviso: Não foi possível fechar o navegador automaticamente: {e}")

def complete_shutdown():
    """Executa shutdown completo da aplicação"""
    global shutdown_initiated
    if shutdown_initiated:
        return
    
    shutdown_initiated = True
    print("\n⏰ Timeout de 3 minutos atingido - Encerrando aplicação...")
    
    try:
        # Limpar uploads
        cleanup_uploads()
        print("✅ Arquivos temporários limpos")
        
        # Fechar navegador (apenas ambiente local/desktop)
        if not IS_CLOUD:
            close_browser()
            print("✅ Navegador fechado")
        
        # Encerrar processo
        print("✅ Encerrando aplicação...")
        os._exit(0)
        
    except Exception as e:
        print(f"Erro durante shutdown: {e}")
        os._exit(1)

def monitor_activity():
    """Thread que monitora a atividade e executa shutdown se necessário"""
    global last_activity_time, shutdown_initiated
    
    while not shutdown_initiated:
        try:
            time.sleep(60)  # Verifica a cada minuto
            
            with activity_lock:
                current_time = time.time()
                inactive_time = current_time - last_activity_time
                
                if inactive_time >= (timeout_minutes * 60):
                    complete_shutdown()
                    break
                    
        except Exception as e:
            print(f"Erro no monitoramento de atividade: {e}")
            time.sleep(60)

def run_app():
    """Executa o servidor Flask com configurações otimizadas"""
    try:
        app.run(
            debug=False,  # Desabilitar debug para produção
            use_reloader=False,
            host='0.0.0.0',
            port=PORT,
            threaded=True  # Permitir múltiplas conexões
        )
    except Exception as e:
        print(f"Erro ao iniciar o servidor Flask: {e}")

def check_server_ready(host='127.0.0.1', port=5000, timeout=15):
    """Verifica se o servidor está pronto para receber conexões"""
    start_time = time.time()
    
    # Primeiro, aguardar a porta estar em uso (servidor iniciado)
    while time.time() - start_time < timeout:
        try:
            with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as s:
                s.settimeout(1)
                result = s.connect_ex((host, port))
                if result == 0:  # Conexão bem-sucedida
                    # Aguardar um pouco mais para o Flask estar completamente pronto
                    time.sleep(2)
                    
                    # Tentar fazer uma requisição HTTP
                    try:
                        response = urllib.request.urlopen(f'http://{host}:{port}/', timeout=3)
                        if response.getcode() == 200:
                            return True
                    except Exception:
                        # Se a requisição HTTP falhar, mas a porta está aberta,
                        # provavelmente o servidor está funcionando
                        return True
        except Exception:
            pass
        
        time.sleep(0.5)
    
    return False

def is_port_available(host='127.0.0.1', port=5000):
    """Verifica se a porta está disponível"""
    try:
        with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as s:
            s.settimeout(1)
            result = s.connect_ex((host, port))
            return result != 0  # Porta disponível se conexão falhar
    except Exception:
        return True

# Função para limpar a pasta uploads
def cleanup_uploads():
    """Remove todos os arquivos da pasta uploads"""
    try:
        upload_folder = app.config['UPLOAD_FOLDER']
        if os.path.exists(upload_folder):
            # Remover todos os arquivos da pasta uploads
            files = glob.glob(os.path.join(upload_folder, '*'))
            for file in files:
                try:
                    if os.path.isfile(file):
                        os.remove(file)
                        print(f"Arquivo removido: {file}")
                except Exception as e:
                    print(f"Erro ao remover arquivo {file}: {e}")
            print("✅ Pasta uploads limpa com sucesso!")
    except Exception as e:
        print(f"Erro ao limpar pasta uploads: {e}")

# Handler para sinais de interrupção
def signal_handler(signum, frame):
    """Handler para capturar sinais de encerramento"""
    print("\n🛑 Encerrando aplicação...")
    cleanup_uploads()
    print("✅ Aplicação encerrada com sucesso!")
    os._exit(0)

if __name__ == '__main__':
    print("=== Conversor Datalogger ===")
    print("Iniciando aplicação...")
    
    # Registrar handlers de limpeza
    signal.signal(signal.SIGINT, signal_handler)  # Ctrl+C
    signal.signal(signal.SIGTERM, signal_handler)  # Terminação
    atexit.register(cleanup_uploads)  # Encerramento normal
    
    # Verificar se a porta está disponível
    if not is_port_available():
        print("Aviso: A porta 5000 já está em uso. Tentando continuar...")
    
    # Iniciar o servidor Flask em thread separada
    print("Iniciando servidor Flask...")
    server_thread = Thread(target=run_app)
    server_thread.daemon = True
    server_thread.start()

    # Iniciar thread de monitoramento de timeout (apenas local)
    if not IS_CLOUD:
        print("Iniciando monitoramento de timeout (3 minutos)...")
        monitor_thread = Thread(target=monitor_activity)
        monitor_thread.daemon = True
        monitor_thread.start()
    
    # Aguardar o servidor estar pronto
    print("Aguardando servidor ficar pronto...")
    
    # Aguardar um tempo fixo para o servidor iniciar
    time.sleep(4)
    
    # Verificar se o servidor está respondendo
    server_ready = check_server_ready()
    
    if server_ready:
        print("✅ Servidor iniciado com sucesso!")
    else:
        print("⚠️ Servidor pode estar iniciando... Tentando abrir navegador mesmo assim.")
    
    if not IS_CLOUD:
        print("🌐 Abrindo navegador...")
        try:
            webbrowser.open(f"http://127.0.0.1:{PORT}/")
            print("✅ Aplicação aberta no navegador!")
        except Exception as e:
            print(f"Erro ao abrir navegador: {e}")
            print(f"Acesse manualmente: http://127.0.0.1:{PORT}/")
    
    if not IS_CLOUD:
        print("\n📋 Instruções:")
        print(f"- A aplicação está rodando em: http://127.0.0.1:{PORT}/")
        print("- Para encerrar, pressione Ctrl+C")
        print("- ⏰ A aplicação será fechada automaticamente após 3 minutos sem atividade")
        print("- Se houver problemas, aguarde alguns segundos e recarregue a página")
        print("- 🗑️ Os arquivos da pasta uploads serão limpos automaticamente ao encerrar")
    
    # Manter o programa rodando
    try:
        while True:
            time.sleep(1)
    except KeyboardInterrupt:
        signal_handler(signal.SIGINT, None)
        signal_handler(signal.SIGINT, None)
